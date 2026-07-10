"""Оборудыш — генератор справочников из Excel для автосверки и автозаполнения.

Режим 1 (по умолчанию): Выписывает ФИО в org_members.csv для автоверификации СО/ССФ.
  python make_members.py "файл.xlsx" [имя листа]

Режим 2 (--directory): Выписывает username;name;deps;role в directory.csv для автозаполнения.
  python make_members.py "файл.xlsx" --directory

Структура Excel для directory: колонки A=username, B=ФИО, C=отделы(через запятую), D=роль
"""
import sys
import csv
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("Нужен openpyxl. Установите:  pip install openpyxl")
    sys.exit(1)


def make_org_members(src: Path, sheet_name: str = None) -> int:
    """Создаёт org_members.csv (только ФИО, по одному на строку)."""
    wb = load_workbook(src, read_only=True, data_only=True)
    sheets = [wb[sheet_name]] if sheet_name else wb.worksheets
    names = set()
    for ws in sheets:
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is None:
                    continue
                s = " ".join(str(cell).split())
                if len(s.split()) >= 2:  # похоже на ФИО
                    names.add(s)
    out = Path(__file__).resolve().parent / "org_members.csv"
    out.write_text("\n".join(sorted(names)), encoding="utf-8")
    return len(names)


def make_directory(src: Path, sheet_name: str = None) -> int:
    """Создаёт directory.csv (username;name;deps;role) для автозаполнения."""
    wb = load_workbook(src, read_only=True, data_only=True)
    sheets = [wb[sheet_name]] if sheet_name else wb.worksheets
    rows = []
    for ws in sheets:
        for row in ws.iter_rows(values_only=True):
            if not row or not row[0]:  # username обязателен
                continue
            username = str(row[0]).strip()
            if not username:
                continue
            name = str(row[1] or "").strip()
            deps = str(row[2] or "").strip()
            role = str(row[3] or "").strip()
            rows.append((username, name, deps, role))
    out = Path(__file__).resolve().parent / "directory.csv"
    with out.open("w", encoding="utf-8", newline="") as f:
        w = csv.writer(f, delimiter=";")
        for r in rows:
            w.writerow(r)
    return len(rows)


def main() -> None:
    if len(sys.argv) < 2:
        print('Использование: python make_members.py "файл.xlsx" [имя листа]')
        print('             python make_members.py "файл.xlsx" --directory [имя листа]')
        sys.exit(1)
    src = Path(sys.argv[1])
    if not src.is_file():
        print("Файл не найден:", src)
        sys.exit(1)
    if "--directory" in sys.argv:
        idx = sys.argv.index("--directory")
        sheet = sys.argv[idx + 1] if idx + 1 < len(sys.argv) else None
        n = make_directory(src, sheet)
        print(f"Готово: {n} записей справочника -> directory.csv")
    else:
        sheet = sys.argv[2] if len(sys.argv) > 2 else None
        n = make_org_members(src, sheet)
        print(f"Готово: {n} имён -> org_members.csv")


if __name__ == "__main__":
    main()